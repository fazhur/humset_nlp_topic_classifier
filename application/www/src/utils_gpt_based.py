from openai import OpenAI
import pandas as pd
import json
from openpyxl import Workbook
import time
from www.src.task_executor import TaskExecutor

list_topics = ['At risk->Risk and vulnerabilities',
 'Capacities & response->International response',
 'Capacities & response->National response',
 'Capacities & response->People reached/response gaps',
 'Casualties->Dead',
 'Casualties->Injured',
 'Context->Demography',
 'Context->Economy',
 'Context->Environment',
 'Context->Legal & policy',
 'Context->Politics',
 'Context->Security & stability',
 'Context->Socio cultural',
 'Covid-19->Cases',
 'Covid-19->Deaths',
 'Covid-19->Prevention campaign',
 'Covid-19->Restriction measures',
 'Covid-19->Testing',
 'Covid-19->Vaccination',
 'Displacement->Push factors',
 'Displacement->Type/numbers/movements',
 'Humanitarian access->Physical constraints',
 'Humanitarian access->Relief to population',
 'Humanitarian conditions->Coping mechanisms',
 'Humanitarian conditions->Living standards',
 'Humanitarian conditions->Number of people in need',
 'Humanitarian conditions->Physical and mental well being',
 'Impact->Driver/aggravating factors',
 'Impact->Impact on people',
 'Impact->Impact on systems, services and networks',
 'Impact->Number of people affected',
 'Information and communication->Communication means and preferences',
 'Information and communication->Knowledge and info gaps (hum)',
 'Information and communication->Knowledge and info gaps (pop)',
 'Priority interventions->Expressed by humanitarian staff',
 'Priority needs->Expressed by humanitarian staff',
 'Priority needs->Expressed by population',
 'Shock/event->Hazard & threats',
 'Shock/event->Type and characteristics',
 'Shock/event->Underlying/aggravating factors'
]


def get_completion(prompt, model = "gpt-4o-mini"):
    
    messages = [{"role": "system", "content": "You are a helpful assistant that detects theses discussed in the text."}, {"role": "user", "content": prompt}]

    client = OpenAI(api_key="")
    response = client.chat.completions.create(
        model=model,
        messages=messages,
        temperature=0,
    )
    return response.choices[0].message.content


def clean_json_output(output):
    response = output.replace("json\n", "").replace("```", "")
    return response


async def create_topics_table(data):
    executor = TaskExecutor()
    result_table = pd.DataFrame(columns=['file_name', 'text', 'topic'])
    start = time.time()
    
    for idx, row in data.iterrows():
        
        topics_dict = {}
        for topic in list_topics:
            topics_dict[topic.lower()] = topic
        
        topics_representaion = '\n'.join('- ' + topic.lower() for topic in list_topics)
        prompt = f"""
        Given the input below, identify the theses than were discussed in the text.
        Format your response in JSON with "main_points" as the key and a list of theses
        that exactly match the points mentioned in the input without external symbols.

        List of atomic theses:\n{topics_representaion}

        Input text:\n"{row["text"]}"
        """
        executor.add_task(prompt)

    responses = await executor.execute()
    for idx, response in enumerate(responses):

        response = clean_json_output(response)
        
        try:
            response_table_json = json.loads(response)
        except Exception as e:
            continue

        file_name = data.iloc[idx]["file_name"]
        topics = response_table_json["main_points"]

        if len(topics) == 0:
            result_table.loc[len(result_table)] = [file_name, data.iloc[idx]['text'], "NA"]
            continue
        
        for topic in topics:
            
            real_topic = topics_dict.get(topic.lower(), None)
            if real_topic is not None:
                result_table.loc[len(result_table)] = [file_name, data.iloc[idx]['text'], real_topic]

                
    print(time.time() - start)
    return result_table


def create_resut_file(topics_table):
    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    result_sheet = workbook.create_sheet("Results")

    rt = topics_table.copy()
    pivot_table = rt.pivot_table(index='topic', columns='file_name', aggfunc='size', fill_value=0)

    columns = pivot_table.columns.tolist()
    header_row = ["Topic"] + [str(col) for col in columns]
    result_sheet.append(header_row)

    for topic, row_values in pivot_table.iterrows():
        result_sheet.append([topic] + row_values.tolist())

    for col_idx, column_cells in enumerate(result_sheet.columns, start=1):
        if col_idx == 1:
            result_sheet.column_dimensions[column_cells[0].column_letter].width = 50
        else:
            result_sheet.column_dimensions[column_cells[0].column_letter].width = 15

    for row in result_sheet.iter_rows():
        for cell in row:
            result_sheet.row_dimensions[cell.row].height = 20

    return workbook


